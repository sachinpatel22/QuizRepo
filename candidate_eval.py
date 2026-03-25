import json
import os
from pathlib import Path

import pandas as pd
from google import genai
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ----------------------------
# Configuration
# ----------------------------
MODEL_NAME = "gemini-3-flash-preview"

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "."
OUTPUT_DIR = BASE_DIR / "output"

CV_FILE = INPUT_DIR / "Python-sample-resume.pdf"  # change if needed: cv.docx or cv.txt
JOB_DESCRIPTION_FILE = INPUT_DIR / "job_description.txt"
QUESTIONS_FILE = INPUT_DIR / "selected_questions.txt"
ANSWERS_FILE = INPUT_DIR / "answers.txt"

JSON_OUTPUT_FILE = OUTPUT_DIR / "candidate_report.json"
EXCEL_OUTPUT_FILE = OUTPUT_DIR / "candidate_report.xlsx"


# ----------------------------
# Response schema
# ----------------------------
RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "executive_summary": {
            "type": "object",
            "properties": {
                "current_role": {"type": "string"},
                "current_role_fit": {"type": "string"},
                "recommended_role": {"type": "string"},
                "headline": {"type": "string"},
                "reason_summary": {"type": "string"},
                "confidence_score": {"type": "integer"},
            },
            "required": [
                "current_role",
                "current_role_fit",
                "recommended_role",
                "headline",
                "reason_summary",
                "confidence_score",
            ],
        },
        "candidate_profile": {
            "type": "object",
            "properties": {
                "candidate_name": {"type": "string"},
                "summary": {"type": "string"},
                "estimated_experience_level": {"type": "string"},
                "core_skills": {"type": "array", "items": {"type": "string"}},
                "secondary_skills": {"type": "array", "items": {"type": "string"}},
                "domain_experience": {"type": "array", "items": {"type": "string"}},
                "strengths": {"type": "array", "items": {"type": "string"}},
                "risks": {"type": "array", "items": {"type": "string"}},
            },
            "required": [
                "candidate_name",
                "summary",
                "estimated_experience_level",
                "core_skills",
                "secondary_skills",
                "domain_experience",
                "strengths",
                "risks",
            ],
        },
        "requirement_fit": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "requirement_type": {"type": "string"},
                    "requirement": {"type": "string"},
                    "fit_level": {"type": "string"},
                    "evidence_from_cv": {"type": "array", "items": {"type": "string"}},
                    "evidence_from_answers": {"type": "array", "items": {"type": "string"}},
                    "gap": {"type": "string"},
                },
                "required": [
                    "requirement_type",
                    "requirement",
                    "fit_level",
                    "evidence_from_cv",
                    "evidence_from_answers",
                    "gap",
                ],
            },
        },
        "question_evaluation": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "question": {"type": "string"},
                    "candidate_answer": {"type": "string"},
                    "expected_topic": {"type": "string"},
                    "rating": {"type": "string"},
                    "score": {"type": "integer"},
                    "strengths": {"type": "array", "items": {"type": "string"}},
                    "concerns": {"type": "array", "items": {"type": "string"}},
                    "missing_points": {"type": "array", "items": {"type": "string"}},
                },
                "required": [
                    "question",
                    "candidate_answer",
                    "expected_topic",
                    "rating",
                    "score",
                    "strengths",
                    "concerns",
                    "missing_points",
                ],
            },
        },
        "final_recommendation": {
            "type": "object",
            "properties": {
                "decision": {"type": "string"},
                "manager_recommendation": {"type": "string"},
                "overall_score": {"type": "integer"},
                "follow_up_needed": {"type": "boolean"},
                "follow_up_areas": {"type": "array", "items": {"type": "string"}},
            },
            "required": [
                "decision",
                "manager_recommendation",
                "overall_score",
                "follow_up_needed",
                "follow_up_areas",
            ],
        },
        "missing_information": {
            "type": "array",
            "items": {"type": "string"},
        },
    },
    "required": [
        "executive_summary",
        "candidate_profile",
        "requirement_fit",
        "question_evaluation",
        "final_recommendation",
        "missing_information",
    ],
}


# ----------------------------
# Prompt
# ----------------------------
PROMPT = """
You are an AI hiring evaluation assistant.

You will receive 4 files:
1. Candidate CV
2. Job Description
3. Selected Questions
4. Candidate Answers

Your task:
Create a comprehensive candidate profile and a structured evaluation of the candidate's fit for the role.

Important decision rule:
- Determine whether the candidate is suitable for the CURRENT role described in the job description.
- If the candidate is not suitable for the current role but appears strong for another role or level, recommend that role explicitly.
- Make the report useful for a hiring manager.

Rules:
- Use only evidence found in the uploaded files.
- Do not invent experience, leadership, projects, or skills.
- If information is missing, say so clearly.
- Extract must-have and nice-to-have requirements from the job description.
- Evaluate how well the CV and answers support each requirement.
- Evaluate each selected question and answer.
- Be professional, specific, and evidence-based.

Output expectations:
- executive_summary.current_role_fit should be one of:
  suitable, borderline, not_suitable
- requirement_fit.fit_level should be one of:
  strong_match, partial_match, weak_match, no_evidence
- question_evaluation.rating should be one of:
  strong, acceptable, weak, incorrect
- final_recommendation.decision should be one of:
  hire_for_current_role, do_not_hire_for_current_role, consider_for_alternative_role, hold_for_more_interview

Write the executive summary so it is easy for a hiring manager to read quickly.
Return JSON only.
""".strip()


# ----------------------------
# Helpers
# ----------------------------
def list_to_text(value):
    if isinstance(value, list):
        return " | ".join(str(x) for x in value)
    return value


def ensure_inputs_exist():
    required_files = [CV_FILE, JOB_DESCRIPTION_FILE, QUESTIONS_FILE, ANSWERS_FILE]
    missing = [str(f) for f in required_files if not f.exists()]
    if missing:
        raise FileNotFoundError(
            "Missing required input file(s):\n" + "\n".join(missing)
        )


def call_gemini():
    api_key = "API-Key" #os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise EnvironmentError("GEMINI_API_KEY is not set in your environment.")

    client = genai.Client(api_key=api_key)

    uploaded_cv = client.files.upload(file=str(CV_FILE))
    uploaded_jd = client.files.upload(file=str(JOB_DESCRIPTION_FILE))
    uploaded_questions = client.files.upload(file=str(QUESTIONS_FILE))
    uploaded_answers = client.files.upload(file=str(ANSWERS_FILE))

    response = client.models.generate_content(
        model=MODEL_NAME,
        contents=[
            uploaded_cv,
            uploaded_jd,
            uploaded_questions,
            uploaded_answers,
            PROMPT,
        ],
        config={
            "response_mime_type": "application/json",
            "response_json_schema": RESPONSE_SCHEMA,
        },
    )

    text = response.text.strip()
    return json.loads(text)


def save_json(data):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(JSON_OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def build_dataframes(data):
    executive = data.get("executive_summary", {})
    profile = data.get("candidate_profile", {})
    final_rec = data.get("final_recommendation", {})

    executive_df = pd.DataFrame([{
        "Current Role": executive.get("current_role", ""),
        "Current Role Fit": executive.get("current_role_fit", ""),
        "Recommended Role": executive.get("recommended_role", ""),
        "Headline": executive.get("headline", ""),
        "Reason Summary": executive.get("reason_summary", ""),
        "Confidence Score": executive.get("confidence_score", ""),
        "Decision": final_rec.get("decision", ""),
        "Overall Score": final_rec.get("overall_score", ""),
        "Manager Recommendation": final_rec.get("manager_recommendation", ""),
        "Follow Up Needed": final_rec.get("follow_up_needed", ""),
        "Follow Up Areas": list_to_text(final_rec.get("follow_up_areas", [])),
        "Missing Information": list_to_text(data.get("missing_information", [])),
    }])

    profile_df = pd.DataFrame([{
        "Candidate Name": profile.get("candidate_name", ""),
        "Summary": profile.get("summary", ""),
        "Estimated Experience Level": profile.get("estimated_experience_level", ""),
        "Core Skills": list_to_text(profile.get("core_skills", [])),
        "Secondary Skills": list_to_text(profile.get("secondary_skills", [])),
        "Domain Experience": list_to_text(profile.get("domain_experience", [])),
        "Strengths": list_to_text(profile.get("strengths", [])),
        "Risks": list_to_text(profile.get("risks", [])),
    }])

    requirement_rows = []
    for item in data.get("requirement_fit", []):
        requirement_rows.append({
            "Requirement Type": item.get("requirement_type", ""),
            "Requirement": item.get("requirement", ""),
            "Fit Level": item.get("fit_level", ""),
            "Evidence from CV": list_to_text(item.get("evidence_from_cv", [])),
            "Evidence from Answers": list_to_text(item.get("evidence_from_answers", [])),
            "Gap": item.get("gap", ""),
        })
    requirement_df = pd.DataFrame(requirement_rows)

    question_rows = []
    for item in data.get("question_evaluation", []):
        question_rows.append({
            "Question": item.get("question", ""),
            "Candidate Answer": item.get("candidate_answer", ""),
            "Expected Topic": item.get("expected_topic", ""),
            "Rating": item.get("rating", ""),
            "Score": item.get("score", ""),
            "Strengths": list_to_text(item.get("strengths", [])),
            "Concerns": list_to_text(item.get("concerns", [])),
            "Missing Points": list_to_text(item.get("missing_points", [])),
        })
    question_df = pd.DataFrame(question_rows)

    return executive_df, profile_df, requirement_df, question_df


def format_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell.alignment = wrap_alignment
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 14), 50)


def save_excel(data):
    executive_df, profile_df, requirement_df, question_df = build_dataframes(data)

    with pd.ExcelWriter(EXCEL_OUTPUT_FILE, engine="openpyxl") as writer:
        executive_df.to_excel(writer, sheet_name="Executive_Summary", index=False)
        profile_df.to_excel(writer, sheet_name="Candidate_Profile", index=False)
        requirement_df.to_excel(writer, sheet_name="Requirement_Fit", index=False)
        question_df.to_excel(writer, sheet_name="Question_Evaluation", index=False)

    wb = load_workbook(EXCEL_OUTPUT_FILE)
    for ws in wb.worksheets:
        format_worksheet(ws)
    wb.save(EXCEL_OUTPUT_FILE)


def main():
    print("Checking input files...")
    ensure_inputs_exist()

    print("Calling Gemini...")
    data = call_gemini()

    print("Saving JSON...")
    save_json(data)

    print("Saving Excel report...")
    save_excel(data)

    print("Done.")
    print(f"JSON  : {JSON_OUTPUT_FILE}")
    print(f"Excel : {EXCEL_OUTPUT_FILE}")


if __name__ == "__main__":
    main()
